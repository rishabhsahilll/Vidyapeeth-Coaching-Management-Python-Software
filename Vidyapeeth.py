# Registration No. | Student Name | Class | Gender | DOB | Date Of Registration	| Address | Phone Number | Father Name | Mother Name | Total Fee Amount	| Amount Paid | Amount Due | Date of Payment

from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import shutil
import os 
from tkinter.ttk import Combobox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib
from tkcalendar import DateEntry
import all_students_data_list

# Data Copy in Other Drive
def data_copy(dst_path="D:\\"):
    # Define the source and destination paths
    src_path = "Student Data"
    # dst_path = "D:\\"

    # Create the destination folder if it does not exist
    if not os.path.exists(src_path):
        try:
            shutil.copytree(f"{dst_path}\\{src_path}", f"{os.getcwd()}\\{src_path}")
        except:
            os.makedirs(f"{src_path}\\Student Images")
            os.makedirs(f"{src_path}\\Student Images")
    elif not os.path.exists(src_path):
        os.makedirs(f"{src_path}\\Student Images")
    try:
        shutil.copytree(src_path, f"{dst_path}\\{src_path}")
    except:
        try:
            shutil.rmtree(f"{dst_path}\\{src_path}")
            shutil.copytree(src_path, f"{dst_path}\\{src_path}")
        except:
            try:
                shutil.copytree(src_path, f"{dst_path}\\{src_path}")
            except:
                pass

# data_copy(dst_path="D:\\")
data_copy(dst_path="C:\\")

background = "#06283D"
framebg = "#EDEDED"
framefg = "#06283D"

root = Tk()

today = date.today()
f1 = today.strftime("%d/%m/%Y")
paydate = "DD-MM-YYY"

root.title("वंदे मातरम् विद्यापीठ | R.S (RSHABH-SAHIL)")
root.iconbitmap("icon.ico")
root.geometry("1250x700+210+100")
root.config(bg=background)


all_student_data = []

file = pathlib.Path("Student Data/Student_Data.xlsx")

if file.exists():
    pass
else:
    try:
        shutil.copytree(f"D:\\Student Data", f"{os.getcwd()}\\Student Data")
        if file.exists():
            pass
        else:
            file = Workbook()
            sheet=file.active
            sheet['A1']="Registration No."
            sheet['B1']="Student Name"
            sheet['C1']="Class"
            sheet['D1']="Gender"
            sheet['E1']="DOB"
            sheet['F1']="Date Of Registration"
            sheet['G1']="Address"
            sheet['H1']="Phone Number"
            sheet['I1']="Father Name"
            sheet['J1']="Mother Name"
            sheet['K1']="Total Fee Amount"
            sheet["L1"]="January"
            sheet["M1"]="February"
            sheet["N1"]="March"
            sheet["O1"]="April"
            sheet["P1"]="May"
            sheet["Q1"]="June"
            sheet["R1"]="July"
            sheet["S1"]="August"
            sheet["T1"]="September"
            sheet["U1"]="October"
            sheet["V1"]="November"
            sheet["W1"]="December"

            file.save("Student Data/Student_Data.xlsx")
    except:
        file = Workbook()
        sheet=file.active
        sheet['A1']="Registration No."
        sheet['B1']="Student Name"
        sheet['C1']="Class"
        sheet['D1']="Gender"
        sheet['E1']="DOB"
        sheet['F1']="Date Of Registration"
        sheet['G1']="Address"
        sheet['H1']="Phone Number"
        sheet['I1']="Father Name"
        sheet['J1']="Mother Name"
        sheet['K1']="Total Fee Amount"
        sheet["L1"]="January"
        sheet["M1"]="February"
        sheet["N1"]="March"
        sheet["O1"]="April"
        sheet["P1"]="May"
        sheet["Q1"]="June"
        sheet["R1"]="July"
        sheet["S1"]="August"
        sheet["T1"]="September"
        sheet["U1"]="October"
        sheet["V1"]="November"
        sheet["W1"]="December"

        file.save("Student Data/Student_Data.xlsx")

# Exit Window
def Exit():
    root.destroy()

# Image Show
def showimage(event=None):
    global filename
    global img

    filename = filedialog.askopenfilename(initialdir=os.getcwd(),title="Select image file",filetype=(("JPG File","*.jpg"),
    ("PNG File","*.png"),
    ("All File","*.*")))
    try:
        img = (Image.open(filename))
        resized_image = img.resize((198,198))
        photo2 = ImageTk.PhotoImage(resized_image)
        lbl.config(image=photo2)
        lbl.image=photo2
    except:
        pass

# Registration No.
# it is created to automatic enter registration no.
def registration_no():
    file = openpyxl.load_workbook("Student Data/Student_Data.xlsx")
    sheet = file.active
    row=sheet.max_row

    max_row_value = sheet.cell(row=row,column=1).value

    try:
        Registration.set(max_row_value+1)
    except:
        Registration.set(1)

# Clear
def Clear():
    global img
    Name.set("")
    DOB.set("")
    Class.set("Select Class")
    Phone.set("91+")
    F_Name.set("")
    M_Name.set("")
    Address.set("")
    T_Fee_Amount.set("")
    Date.set(f1)
    Search.set("")
    # Amount_Due.set("")
    # Amount_Paid.set("")
    # radio.set(1)
    registration_no()
    saveButton.config(state="normal")
    img1 = PhotoImage(file="Images/student.png")
    lbl.config(image=img1)
    lbl.image = img1
    img = ""
    T_Fee_Amount.set("₹ ")
    Search_Result_List.set("All Student List")
    # jan.set("Payment")
    january_month.set("Payment")
    february_month.set("Payment")
    march_month.set("Payment")
    april_month.set("Payment")
    may_month.set("Payment")
    June_month.set("Payment")
    July_month.set("Payment")
    August_month.set("Payment")
    September_month.set("Payment")
    October_month.set("Payment")
    November_month.set("Payment")
    December_month.set("Payment")

    Janrs.set("₹ ")
    Febrs.set("₹ ")
    Marrs.set("₹ ")
    aprrs.set("₹ ")
    mayrs.set("₹ ")
    juners.set("₹ ")
    julrs.set("₹ ")
    augrs.set("₹ ")
    seprs.set("₹ ")
    octrs.set("₹ ")
    novrs.set("₹ ")
    decrs.set("₹ ")

    JanDate.set(paydate)
    FebDate.set(paydate)
    MarDate.set(paydate)
    ApDate.set(paydate)
    MayDate.set(paydate)
    JunDate.set(paydate)
    JulyDate.set(paydate)
    AugustDate.set(paydate)
    SeptemberDate.set(paydate)
    OctoberDate.set(paydate)
    NovemberDate.set(paydate)
    DecemberDate.set(paydate)
    data_copy(dst_path="D:\\")
    data_copy(dst_path="C:\\")

# Clear
def Clear_Search():
    global img
    Name.set("")
    DOB.set("")
    Class.set("Select Class")
    Phone.set("91+")
    F_Name.set("")
    M_Name.set("")
    Address.set("")
    T_Fee_Amount.set("")
    Date.set(f1)
    Search.set("")
    # Amount_Due.set("")
    # Amount_Paid.set("")
    # radio.set(1)
    registration_no()
    saveButton.config(state="normal")
    img1 = PhotoImage(file="Images/student.png")
    lbl.config(image=img1)
    lbl.image = img1
    img = ""
    T_Fee_Amount.set("₹ ")
    # Search_Result_List.set("All Student List")
    # jan.set("Payment")
    january_month.set("Payment")
    february_month.set("Payment")
    march_month.set("Payment")
    april_month.set("Payment")
    may_month.set("Payment")
    June_month.set("Payment")
    July_month.set("Payment")
    August_month.set("Payment")
    September_month.set("Payment")
    October_month.set("Payment")
    November_month.set("Payment")
    December_month.set("Payment")

    Janrs.set("₹ ")
    Febrs.set("₹ ")
    Marrs.set("₹ ")
    aprrs.set("₹ ")
    mayrs.set("₹ ")
    juners.set("₹ ")
    julrs.set("₹ ")
    augrs.set("₹ ")
    seprs.set("₹ ")
    octrs.set("₹ ")
    novrs.set("₹ ")
    decrs.set("₹ ")

    JanDate.set(paydate)
    FebDate.set(paydate)
    MarDate.set(paydate)
    ApDate.set(paydate)
    MayDate.set(paydate)
    JunDate.set(paydate)
    JulyDate.set(paydate)
    AugustDate.set(paydate)
    SeptemberDate.set(paydate)
    OctoberDate.set(paydate)
    NovemberDate.set(paydate)
    DecemberDate.set(paydate)
    data_copy(dst_path="D:\\")
    data_copy(dst_path="C:\\")


# Save
def Save():
    R1 = Registration.get()
    N1 = Name.get()
    C1 = Class.get()
    D1 = DOB.get()
    P1 = f"91+{str(Phone.get()).replace(' ','').replace('91+','')}"
    FN1 = F_Name.get()
    MN1 = M_Name.get()
    AD1 = Address.get()
    # TF1 = f"₹ {str(T_Fee_Amount.get()).replace(' ','').replace('₹','')}"
    TF1 = f"₹ {str(T_Fee_Amount.get()).replace(' ','').replace('₹','')}"
    # AMD1 = Amount_Due.get()
    # AMP1 = Amount_Paid.get()
    DR1 = Date.get()
    # DOPM1 = Date_of_Payment.get()
    jan1 = january_month.get()
    jan2 = february_month.get()
    jan3 = march_month.get()
    jan4 = april_month.get()
    jan5 = may_month.get()
    jan6 = June_month.get()
    jan7 = July_month.get()
    jan8 = August_month.get()
    jan9 = September_month.get()
    jan10 = October_month.get()
    jan11 = November_month.get()
    jan12 = December_month.get()

    jan_rs1 =  f"₹ {str(Janrs.get()).replace(' ','').replace('₹','')}"
    jan_rs2 =  f"₹ {str(Febrs.get()).replace(' ','').replace('₹','')}"
    jan_rs3 =  f"₹ {str(Marrs.get()).replace(' ','').replace('₹','')}"
    jan_rs4 =  f"₹ {str(aprrs.get()).replace(' ','').replace('₹','')}"    
    jan_rs5 =  f"₹ {str(mayrs.get()).replace(' ','').replace('₹','')}"
    jan_rs6 =  f"₹ {str(juners.get()).replace(' ','').replace('₹','')}"
    jan_rs7 =  f"₹ {str(julrs.get()).replace(' ','').replace('₹','')}"
    jan_rs8 =  f"₹ {str(augrs.get()).replace(' ','').replace('₹','')}"
    jan_rs9 =  f"₹ {str(seprs.get()).replace(' ','').replace('₹','')}"
    jan_rs10 = f"₹ {str(octrs.get()).replace(' ','').replace('₹','')}"
    jan_rs11 = f"₹ {str(novrs.get()).replace(' ','').replace('₹','')}"
    jan_rs12 = f"₹ {str(decrs.get()).replace(' ','').replace('₹','')}"

    jan_date1 = JanDate.get()
    jan_date2 = FebDate.get()
    jan_date3 = MarDate.get()
    jan_date4 = ApDate.get()
    jan_date5 = MayDate.get()
    jan_date6 = JunDate.get()
    jan_date7 = JulyDate.get()
    jan_date8 = AugustDate.get()
    jan_date9 = SeptemberDate.get()
    jan_date10 = OctoberDate.get()
    jan_date11 = NovemberDate.get()
    jan_date12 = DecemberDate.get()

    try:
        G1 = gender 
    except:
        messagebox.showerror("error","Select Gender")
    
    if R1=="" or C1=="Select Class" or N1=="" or D1=="" or FN1=="" or MN1=="" or AD1=="" or TF1=="" or P1=="" or P1=="91+" or P1=="91+ ":
        messagebox.showerror("error","Few Data is missing!")

    else:
        try:
            try:
                if int(len(str(P1)))>13:
                    messagebox.showerror("error","Your Phone is Wrong!")
            except:
                messagebox.showerror("error","Please Correct Phone Number!")

            # try:
            #     aa = int(str(jan_rs1).replace(".","0")) + 1
            #     aa = int(str(jan_rs2).replace(".","0")) + 1
            #     aa = int(str(jan_rs3).replace(".","0")) + 1
            #     aa = int(str(jan_rs4).replace(".","0")) + 1
            #     aa = int(str(jan_rs5).replace(".","0")) + 1
            #     aa = int(str(jan_rs6).replace(".","0")) + 1
            #     aa = int(str(jan_rs7).replace(".","0")) + 1
            #     aa = int(str(jan_rs8).replace(".","0")) + 1
            #     aa = int(str(jan_rs9).replace(".","0")) + 1
            #     aa = int(str(jan_rs10).replace(".","0")) + 1
            #     aa = int(str(jan_rs11).replace(".","0")) + 1
            #     aa = int(str(jan_rs12).replace(".","0")) + 1
            #     aa = int(str(TF1).replace(".","0")) + 1
            # except:
            #     messagebox.showerror("error","Please Enter Pyment Only Number!")
            #     return

            file = openpyxl.load_workbook("Student Data/Student_Data.xlsx")
            sheet = file.active

            # if str(P1)[0:2] == "91+" or int(len(str(P1).replace("91+","")))==10:
            #     sheet.cell(column=8,row=sheet.max_row,value=P1)
                
            # elif str(P1)[0:2]!="91+" or int(len(str(P1)))==10:
            #     sheet.cell(column=8,row=sheet.max_row,value=f"91+{P1}")

            # if str(P1.find("91+"))=="0" and int(len(P1))==13 and str(TF1.find("₹"))=="0" and str(jan_rs1.find("₹"))=="0" and str(jan_rs2.find("₹"))=="0" and str(jan_rs3.find("₹"))=="0" and str(jan_rs4.find("₹"))=="0" and str(jan_rs5.find("₹"))=="0" and str(jan_rs6.find("₹"))=="0" and str(jan_rs7.find("₹"))=="0" and str(jan_rs8.find("₹"))=="0" and str(jan_rs9.find("₹"))=="0" and str(jan_rs10.find("₹"))=="0" and str(jan_rs11.find("₹"))=="0" and str(jan_rs12.find("₹"))=="0":
            sheet.cell(column=1,row=sheet.max_row+1,value=R1)
            sheet.cell(column=2,row=sheet.max_row,value=N1)
            sheet.cell(column=3,row=sheet.max_row,value=C1)
            sheet.cell(column=4,row=sheet.max_row,value=G1)
            sheet.cell(column=5,row=sheet.max_row,value=D1)
            sheet.cell(column=6,row=sheet.max_row,value=DR1)
            sheet.cell(column=7,row=sheet.max_row,value=AD1)
            sheet.cell(column=8,row=sheet.max_row,value=P1)

            # sheet.cell(column=8,row=sheet.max_row,value=P1)
            sheet.cell(column=9,row=sheet.max_row,value=FN1)
            sheet.cell(column=10,row=sheet.max_row,value=MN1)

            sheet.cell(column=11,row=sheet.max_row,value=TF1)
            # sheet.cell(column=12,row=sheet.max_row,value=int(AMP1))
            # sheet.cell(column=13,row=sheet.max_row,value=int(AMD1))
            # sheet.cell(column=14,row=sheet.max_row,value=DOPM1)
            sheet.cell(column=12,row=sheet.max_row,value=jan1+" , "+str(jan_rs1)+" , "+jan_date1)
            sheet.cell(column=13,row=sheet.max_row,value=jan2+" , "+str(jan_rs2)+" , "+jan_date2)
            sheet.cell(column=14,row=sheet.max_row,value=jan3+" , "+str(jan_rs3)+" , "+jan_date3)
            sheet.cell(column=15,row=sheet.max_row,value=jan4+" , "+str(jan_rs4)+" , "+jan_date4)
            sheet.cell(column=16,row=sheet.max_row,value=jan5+" , "+str(jan_rs5)+" , "+jan_date5)
            sheet.cell(column=17,row=sheet.max_row,value=jan6+" , "+str(jan_rs6)+" , "+jan_date6)
            sheet.cell(column=18,row=sheet.max_row,value=jan7+" , "+str(jan_rs7)+" , "+jan_date7)
            sheet.cell(column=19,row=sheet.max_row,value=jan8+" , "+str(jan_rs8)+" , "+jan_date8)
            sheet.cell(column=20,row=sheet.max_row,value=jan9+" , "+str(jan_rs9)+" , "+jan_date9)
            sheet.cell(column=21,row=sheet.max_row,value=jan10+" , "+str(jan_rs10)+" , "+jan_date10)
            sheet.cell(column=22,row=sheet.max_row,value=jan11+" , "+str(jan_rs11)+" , "+jan_date11)
            sheet.cell(column=23,row=sheet.max_row,value=jan12+" , "+str(jan_rs12)+" , "+jan_date12)

            file.save(r"Student Data/Student_Data.xlsx")
            try:
                img.save("Student Data/Student Images/"+str(R1)+".jpg")
            except:
                messagebox.showinfo("info",'Profile Picture is not available!!!')

            messagebox.showinfo('Success', 'Successfully data entered!!!')
            root.update()
            
            # x1 = sheet.cell(row=int(R1),column=1).value
            # x2 = sheet.cell(row=int(R1),column=2).value
            # x3 = sheet.cell(row=int(R1),column=3).value
            # all_student_data.append(str(f"{x1}.> {x2} Class: {x3}"))
            # list_data()

            file = openpyxl.load_workbook("Student Data/Student_Data.xlsx")
            sheet = file.active

            Clear()
            registration_no()
            data_copy(dst_path="D:\\")
            data_copy(dst_path="C:\\")

        except:
            messagebox.showerror("error","Please Fill Correct Data!!")
    # print(all_students_data_list.all_student_data)
    # All_Data_Search
    # all_students_data_list.ref_data()
    # Search_Result_List = Combobox(root,values=all_students_data_list.all_student_data,font="arial 13 bold",width=20,height=2,state="r")
    # Search_Result_List.place(x=500,y=150,height=38)
    # Search_Result_List.set("All Student List")
    data_copy(dst_path="D:\\")
    data_copy(dst_path="C:\\")

# Search
def search():
    # print(all_students_data_list.all_student_data.append("Saved"))
    text = Search.get()
    # if text=="" or text==" ":
    #     pass
    # else:
    # print(all_students_data_list.search_data_in_option(query=text))
    # print(all_students_data_list.all_student_data)
    Clear_Search()
    saveButton.config(state='disable')
    file = openpyxl.load_workbook("Student Data/Student_Data.xlsx")
    sheet = file.active

    for row in sheet.rows:
        if str(row[0].value)==str(text):
            name=row[0]
            # print(str(name))
            reg_no_position = str(name)[14:-1]
            reg_number = str(name)[15:-1]
            x1 = sheet.cell(row=int(reg_number),column=1).value
            x2 = sheet.cell(row=int(reg_number),column=2).value
            x3 = sheet.cell(row=int(reg_number),column=3).value
            all_student_data.append(str(f"{x1}.> {x2} Class: {x3}"))
            # print(reg_no_position)
            # print(reg_number)

        elif str(row[7].value).lower() ==str(text).lower() or str(row[7].value).lower()==str("91+"+str(text)).lower():
            name=row[0]
            # print(str(name))
            reg_no_position = str(name)[14:-1]
            reg_number = str(name)[15:-1]
            x1 = sheet.cell(row=int(reg_number),column=1).value
            x2 = sheet.cell(row=int(reg_number),column=2).value
            x3 = sheet.cell(row=int(reg_number),column=3).value
            all_student_data.append(str(f"{x1}.> {x2} Class: {x3}"))
            # print(reg_no_position)
            # print(reg_number)

        elif str(text).lower() in str(row[1].value).lower():
            name=row[0]
            # print(str(name))
            reg_no_position = str(name)[14:-1]
            reg_number = str(name)[15:-1]
            x1 = sheet.cell(row=int(reg_number),column=1).value
            x2 = sheet.cell(row=int(reg_number),column=2).value
            x3 = sheet.cell(row=int(reg_number),column=3).value
            all_student_data.append(str(f"{x1}.> {x2} Class: {x3}"))
            # print(f"{x1}.> {x2} Class: {x3}")
            # print(reg_no_position)
            # print(reg_number)

    try:
        # print(str(name))
        str(name)
    except:
        messagebox.showerror("Invalid","Invalid Your Input!")

    # reg_no_position showing like A2,A3,A4......,AX
    # reg_number just showing number A2 like 2,3,........,x

    try:
        x1 = sheet.cell(row=int(reg_number),column=1).value
        x2 = sheet.cell(row=int(reg_number),column=2).value
        x3 = sheet.cell(row=int(reg_number),column=3).value
        x4 = sheet.cell(row=int(reg_number),column=4).value
        x5 = sheet.cell(row=int(reg_number),column=5).value
        x6 = sheet.cell(row=int(reg_number),column=6).value
        x7 = sheet.cell(row=int(reg_number),column=7).value
        x8 = sheet.cell(row=int(reg_number),column=8).value
        x9 = sheet.cell(row=int(reg_number),column=9).value
        x10 = sheet.cell(row=int(reg_number),column=10).value
        x11 = sheet.cell(row=int(reg_number),column=11).value
        x12 = sheet.cell(row=int(reg_number),column=12).value
        x13 = sheet.cell(row=int(reg_number),column=13).value
        x14 = sheet.cell(row=int(reg_number),column=14).value
        x15 = sheet.cell(row=int(reg_number),column=15).value
        x16 = sheet.cell(row=int(reg_number),column=16).value
        x17 = sheet.cell(row=int(reg_number),column=17).value
        x18 = sheet.cell(row=int(reg_number),column=18).value
        x19 = sheet.cell(row=int(reg_number),column=19).value
        x20 = sheet.cell(row=int(reg_number),column=20).value
        x21 = sheet.cell(row=int(reg_number),column=21).value
        x22 = sheet.cell(row=int(reg_number),column=22).value
        x23 = sheet.cell(row=int(reg_number),column=23).value

        Registration.set(x1)
        Name.set(x2)
        DOB.set(x5)
        Class.set(x3)
        Phone.set(x8)
        F_Name.set(x9)
        M_Name.set(x10)
        Address.set(x7)
        T_Fee_Amount.set(x11)
        january_month.set(x12.split(", ")[0])
        february_month.set(x13.split(", ")[0])
        march_month.set(x14.split(", ")[0])
        april_month.set(x15.split(", ")[0])
        may_month.set(x16.split(", ")[0])
        June_month.set(x17.split(", ")[0])
        July_month.set(x18.split(", ")[0])
        August_month.set(x19.split(", ")[0])
        September_month.set(x20.split(", ")[0])
        October_month.set(x21.split(", ")[0])
        November_month.set(x22.split(", ")[0])
        December_month.set(x23.split(", ")[0])

        Janrs.set(x12.split(", ")[1])
        Febrs.set(x13.split(", ")[1])
        Marrs.set(x14.split(", ")[1])
        aprrs.set(x15.split(", ")[1])
        mayrs.set(x16.split(", ")[1])
        juners.set(x17.split(", ")[1])
        julrs.set(x18.split(", ")[1])
        augrs.set(x19.split(", ")[1])
        seprs.set(x20.split(", ")[1])
        octrs.set(x21.split(", ")[1])
        novrs.set(x22.split(", ")[1])
        decrs.set(x23.split(", ")[1])

        JanDate.set(x12.split(", ")[2])
        FebDate.set(x13.split(", ")[2])
        MarDate.set(x14.split(", ")[2])
        ApDate.set(x15.split(", ")[2])
        MayDate.set(x16.split(", ")[2])
        JunDate.set(x17.split(", ")[2])
        JulyDate.set(x18.split(", ")[2])
        AugustDate.set(x19.split(", ")[2])
        SeptemberDate.set(x20.split(", ")[2])
        OctoberDate.set(x21.split(", ")[2])
        NovemberDate.set(x22.split(", ")[2])
        DecemberDate.set(x23.split(", ")[2])
        
        # Amount_Due.set(x13)
        # Amount_Paid.set(x12)
        Date.set(x6)
        # Date_of_Payment.set(x14)
        Search.set("")
        # print(all_students_data_list.search_data_in_option(query=text))
        # print(all_students_data_list.all_student_data)
        # # all_students_data_list.search_data_in_option(text)
        # Search_Result_List = Combobox(root,values=all_students_data_list.all_student_data,font="arial 13 bold",width=20,height=2,state="r")
        # Search_Result_List.place(x=500,y=150,height=38)
        # Search_Result_List.set(str(f"{int(x1)}.> {x2} Class: {x3}"))
        
        if x4=='Female':
            R2.select()
        else:
            R1.select()
        # Search_Result_List.set(text)
        try:
            img = (Image.open("Student Data/Student Images/"+str(x1)+".jpg"))
            resized_image = img.resize((190,190))
            photo2 = ImageTk.PhotoImage(resized_image)
            lbl.config(image=photo2)
            lbl.image = photo2
        except:
            pass
    except:
        pass

    if str(text)=="":
        SRL1 = Search_Result_List.get()
        print(SRL1)
        try:
            if SRL1=="All Student List":
                pass
            else:
                reg_number = str(int(SRL1[:SRL1.find(".>")])+1)
        except:
            pass
        print("Reg:- "+reg_number)
        try:
            if reg_number==str(0):
                pass
            else:
                try:
                    x1 = sheet.cell(row=int(reg_number),column=1).value
                    x2 = sheet.cell(row=int(reg_number),column=2).value
                    x3 = sheet.cell(row=int(reg_number),column=3).value
                    x4 = sheet.cell(row=int(reg_number),column=4).value
                    x5 = sheet.cell(row=int(reg_number),column=5).value
                    x6 = sheet.cell(row=int(reg_number),column=6).value
                    x7 = sheet.cell(row=int(reg_number),column=7).value
                    x8 = sheet.cell(row=int(reg_number),column=8).value
                    x9 = sheet.cell(row=int(reg_number),column=9).value
                    x10 = sheet.cell(row=int(reg_number),column=10).value
                    x11 = sheet.cell(row=int(reg_number),column=11).value
                    x12 = sheet.cell(row=int(reg_number),column=12).value
                    x13 = sheet.cell(row=int(reg_number),column=13).value
                    x14 = sheet.cell(row=int(reg_number),column=14).value
                    x15 = sheet.cell(row=int(reg_number),column=15).value
                    x16 = sheet.cell(row=int(reg_number),column=16).value
                    x17 = sheet.cell(row=int(reg_number),column=17).value
                    x18 = sheet.cell(row=int(reg_number),column=18).value
                    x19 = sheet.cell(row=int(reg_number),column=19).value
                    x20 = sheet.cell(row=int(reg_number),column=20).value
                    x21 = sheet.cell(row=int(reg_number),column=21).value
                    x22 = sheet.cell(row=int(reg_number),column=22).value
                    x23 = sheet.cell(row=int(reg_number),column=23).value


                    Registration.set(x1)
                    Name.set(x2)
                    DOB.set(x5)
                    Class.set(x3)
                    Phone.set(x8)
                    F_Name.set(x9)
                    M_Name.set(x10)
                    Address.set(x7)
                    T_Fee_Amount.set(x11)

                    january_month.set(x12.split(", ")[0])
                    february_month.set(x13.split(", ")[0])
                    march_month.set(x14.split(", ")[0])
                    april_month.set(x15.split(", ")[0])
                    may_month.set(x16.split(", ")[0])
                    June_month.set(x17.split(", ")[0])
                    July_month.set(x18.split(", ")[0])
                    August_month.set(x19.split(", ")[0])
                    September_month.set(x20.split(", ")[0])
                    October_month.set(x21.split(", ")[0])
                    November_month.set(x22.split(", ")[0])
                    December_month.set(x23.split(", ")[0])

                    Janrs.set(x12.split(", ")[1])
                    Febrs.set(x13.split(", ")[1])
                    Marrs.set(x14.split(", ")[1])
                    aprrs.set(x15.split(", ")[1])
                    mayrs.set(x16.split(", ")[1])
                    juners.set(x17.split(", ")[1])
                    julrs.set(x18.split(", ")[1])
                    augrs.set(x19.split(", ")[1])
                    seprs.set(x20.split(", ")[1])
                    octrs.set(x21.split(", ")[1])
                    novrs.set(x22.split(", ")[1])
                    decrs.set(x23.split(", ")[1])

                    JanDate.set(x12.split(", ")[2])
                    FebDate.set(x13.split(", ")[2])
                    MarDate.set(x14.split(", ")[2])
                    ApDate.set(x15.split(", ")[2])
                    MayDate.set(x16.split(", ")[2])
                    JunDate.set(x17.split(", ")[2])
                    JulyDate.set(x18.split(", ")[2])
                    AugustDate.set(x19.split(", ")[2])
                    SeptemberDate.set(x20.split(", ")[2])
                    OctoberDate.set(x21.split(", ")[2])
                    NovemberDate.set(x22.split(", ")[2])
                    DecemberDate.set(x23.split(", ")[2])
                    # Amount_Due.set(x13)
                    # Amount_Due.set(x13)
                    # Amount_Paid.set(x12)
                    Date.set(x6)
                    # Date_of_Payment.set(x14)
                    Search.set("")
                    # Search_Result_List.set(str(f"{int(x1)}.> {x2} Class: {x3}"))

                    if x4=='Female':
                        R2.select()
                    else:
                        R1.select()

                    try:
                        img = (Image.open("Student Data/Student Images/"+str(x1)+".jpg"))
                        resized_image = img.resize((190,190))
                        photo2 = ImageTk.PhotoImage(resized_image)
                        lbl.config(image=photo2)
                        lbl.image = photo2
                    except:
                        pass

                except:
                    pass
        except:
            pass

    data_copy(dst_path="D:\\")
    data_copy(dst_path="C:\\")
    # Search_Result_List = Combobox(root,values=all_students_data_list.all_student_data,font="arial 13 bold",width=20,height=2,state="r")
    # Search_Result_List.place(x=500,y=150,height=38)
    # Search_Result_List.set("All Student List")

# Update
def Update():
    R1 = Registration.get()
    N1 = Name.get()
    C1 = Class.get()
    selection()
    G1=gender
    D1 = DOB.get()
    P1 = f"91+{str(Phone.get()).replace(' ','').replace('91+','')}"
    FN1 = F_Name.get()
    MN1 = M_Name.get()
    AD1 = Address.get()
    TF1 =  f"₹ {str(T_Fee_Amount.get()).replace(' ','').replace('₹','')}"
    
    jan1 = january_month.get()
    jan2 = february_month.get()
    jan3 = march_month.get()
    jan4 = april_month.get()
    jan5 = may_month.get()
    jan6 = June_month.get()
    jan7 = July_month.get()
    jan8 = August_month.get()
    jan9 = September_month.get()
    jan10 = October_month.get()
    jan11 = November_month.get()
    jan12 = December_month.get()

    jan_rs1 =  f"₹ {str(Janrs.get()).replace(' ','').replace('₹','')}"
    jan_rs2 =  f"₹ {str(Febrs.get()).replace(' ','').replace('₹','')}"
    jan_rs3 =  f"₹ {str(Marrs.get()).replace(' ','').replace('₹','')}"
    jan_rs4 =  f"₹ {str(aprrs.get()).replace(' ','').replace('₹','')}"    
    jan_rs5 =  f"₹ {str(mayrs.get()).replace(' ','').replace('₹','')}"
    jan_rs6 =  f"₹ {str(juners.get()).replace(' ','').replace('₹','')}"
    jan_rs7 =  f"₹ {str(julrs.get()).replace(' ','').replace('₹','')}"
    jan_rs8 =  f"₹ {str(augrs.get()).replace(' ','').replace('₹','')}"
    jan_rs9 =  f"₹ {str(seprs.get()).replace(' ','').replace('₹','')}"
    jan_rs10 = f"₹ {str(octrs.get()).replace(' ','').replace('₹','')}"
    jan_rs11 = f"₹ {str(novrs.get()).replace(' ','').replace('₹','')}"
    jan_rs12 = f"₹ {str(decrs.get()).replace(' ','').replace('₹','')}"

    jan_date1 = JanDate.get()
    jan_date2 = FebDate.get()
    jan_date3 = MarDate.get()
    jan_date4 = ApDate.get()
    jan_date5 = MayDate.get()
    jan_date6 = JunDate.get()
    jan_date7 = JulyDate.get()
    jan_date8 = AugustDate.get()
    jan_date9 = SeptemberDate.get()
    jan_date10 = OctoberDate.get()
    jan_date11 = NovemberDate.get()
    jan_date12 = DecemberDate.get()
    # AMD1 = Amount_Due.get()
    # AMP1 = Amount_Paid.get()
    DR1 = Date.get()
    # DOPM1 = Date_of_Payment.get()

    file = openpyxl.load_workbook("Student Data/Student_Data.xlsx")
    sheet = file.active

    for row in sheet.rows:
        if row[0].value==R1:
            name = row[0]
            # print(str(name))
            reg_no_position = str(name)[14:-1]
            reg_number = str(name)[15:-1]
            # print(reg_number)
    try:
        # sheet.cell(column=1,row=int(reg_number),value=R1)
        # if str(P1.find("91+"))=="0" and int(len(P1))==13 and str(TF1.find("₹"))=="0" and str(jan_rs1.find("₹"))=="0" and str(jan_rs2.find("₹"))=="0" and str(jan_rs3.find("₹"))=="0" and str(jan_rs4.find("₹"))=="0" and str(jan_rs5.find("₹"))=="0" and str(jan_rs6.find("₹"))=="0" and str(jan_rs7.find("₹"))=="0" and str(jan_rs8.find("₹"))=="0" and str(jan_rs9.find("₹"))=="0" and str(jan_rs10.find("₹"))=="0" and str(jan_rs11.find("₹"))=="0" and str(jan_rs12.find("₹"))=="0":    
        sheet.cell(column=2,row=int(reg_number),value=N1)
        sheet.cell(column=3,row=int(reg_number),value=C1)
        sheet.cell(column=4,row=int(reg_number),value=G1)
        sheet.cell(column=5,row=int(reg_number),value=D1)
        sheet.cell(column=6,row=int(reg_number),value=DR1)
        sheet.cell(column=7,row=int(reg_number),value=AD1)
        sheet.cell(column=8,row=int(reg_number),value=P1)
        sheet.cell(column=9,row=int(reg_number),value=FN1)
        sheet.cell(column=10,row=int(reg_number),value=MN1)
        sheet.cell(column=11,row=int(reg_number),value=TF1)
        # sheet.cell(column=12,row=int(reg_number),value=AMP1)
        # sheet.cell(column=13,row=int(reg_number),value=AMD1)
        # sheet.cell(column=14,row=int(reg_number),value=DOPM1)
        sheet.cell(column=12,row=int(reg_number),value=jan1+" , "+str(jan_rs1)+" , "+jan_date1)
        sheet.cell(column=13,row=int(reg_number),value=jan2+" , "+str(jan_rs2)+" , "+jan_date2)
        sheet.cell(column=14,row=int(reg_number),value=jan3+" , "+str(jan_rs3)+" , "+jan_date3)
        sheet.cell(column=15,row=int(reg_number),value=jan4+" , "+str(jan_rs4)+" , "+jan_date4)
        sheet.cell(column=16,row=int(reg_number),value=jan5+" , "+str(jan_rs5)+" , "+jan_date5)
        sheet.cell(column=17,row=int(reg_number),value=jan6+" , "+str(jan_rs6)+" , "+jan_date6)
        sheet.cell(column=18,row=int(reg_number),value=jan7+" , "+str(jan_rs7)+" , "+jan_date7)
        sheet.cell(column=19,row=int(reg_number),value=jan8+" , "+str(jan_rs8)+" , "+jan_date8)
        sheet.cell(column=20,row=int(reg_number),value=jan9+" , "+str(jan_rs9)+" , "+jan_date9)
        sheet.cell(column=21,row=int(reg_number),value=jan10+" , "+str(jan_rs10)+" , "+jan_date10)
        sheet.cell(column=22,row=int(reg_number),value=jan11+" , "+str(jan_rs11)+" , "+jan_date11)
        sheet.cell(column=23,row=int(reg_number),value=jan12+" , "+str(jan_rs12)+" , "+jan_date12)
        file.save(r"Student Data/Student_Data.xlsx")

        try:
            img.save("Student Data/Student Images/"+str(R1)+".jpg")
        except:
            pass

        messagebox.showinfo("Update","Update Successfully!!!")

        Clear()
        data_copy(dst_path="D:\\")
        data_copy(dst_path="C:\\")

    except:
        pass

    data_copy(dst_path="D:\\")
    data_copy(dst_path="C:\\")

# Gender Selection Function
def selection():
    global gender
    value = radio.get()
    if value==1:
        gender="Male"
        # print(gender)
    else:
        gender="Female"  
        # print(gender)

# month_selection()
# # Add Background Images
bg_image = PhotoImage(file="Images/bg-img.png")
label1 = Label(root, image=bg_image, bg="#FF5800", width=300, height=130)
label1.pack(side=TOP, fill=X)

# # # create another label with the background image and fixed size for the second label
# label2 = Label(root, bg="#f0687c", width=300, height=50)
# label2.pack(side=TOP, fill=X)

# # add the text to the first label
# label1_text = Label(label1, text="R.S (R$HABH-&AHIL)", bg="#f0687c", anchor="e", font=("Arial", 12))
# label1_text.pack(fill=BOTH, expand=1)

# # add the text to the second label
# label2_text = Label(label2, text="वंदे मातरम् विद्यापीठ", bg="#f0687c", fg="#fff", font=("Arial", 20, "bold"))
# label2_text.pack(fill=BOTH, expand=1)

# # create a Label widget and use the image as its background
# bg_label = Label(root, image=bg_image)
# bg_label.place(x=0, y=0, relwidth=1, relheight=1)

# Top Frames
# Label(root,text="वंदे मातरम् विद्यापीठ",width=10,height=2,bg="#f0687c",fg="#fff",font='arial 20 bold').pack(side=TOP,fill=X)
# Label(root,textvariable=Search,width=15,bd=2,font="arial 20").place(x=750,y=150).pack(side=TOP,fill=X)

# Search Box To Update
Search = StringVar()
Entry(root,textvariable=Search,width=10,bd=2,font="arial 20").place(x=720,y=150)
imageicon3 = PhotoImage(file="Images/search.png")
Srch = Button(root,image=imageicon3,width=38,bg="#68ddfa",font="arial 13 bold",command=search)
# Srch = Button(root,text="Search",command=LEFT,width=123,bg="#68ddfa",font="arial 13 bold")
Srch.place(x=888,y=150,height=38)

# imageicon4 = PhotoImage(file="Images/layer 4.png")
# Update_button = Button(root,image=imageicon4,bg="#c36464")
# Update_button = Button(root,text="Update",bg="#c36464",width=10,height=2,font="arial 13 bold",command=Update)
# Update_button.place(x=110,y=64)
# Registration No. | Student Name | Class | Gender | DOB | Date Of Registration	| Address | Phone Number | Father Name | Mother Name | Total Fee Amount	| Amount Paid | Amount Due | Date of Payment

# Registration And Date
Label(root,text="Registration No:",font="arial 13", fg = framebg,bg=background).place(x=30,y=160)
Label(root,text="Date:",font="arial 13", fg = framebg,bg=background).place(x=300,y=160)

Registration = IntVar()
Date = StringVar()

reg_entry = Entry(root,textvariable=Registration,width=15,font="arial 13 bold")
reg_entry.place(x=160,y=150,height=38)

registration_no() # colled it here

# print(f1)
date_entry = Entry(root,textvariable=Date,width=15,font="arial 13 bold")
date_entry.place(x=350,y=150,height=38)

Date.set(f1)

# All_Data_Search
Search_Result_List = Combobox(root,values=all_students_data_list.all_student_data,font="arial 13 bold",width=20,height=2,state="r")
Search_Result_List.place(x=500,y=150,height=38)
Search_Result_List.set("All Student List")

# Student Details
obj= LabelFrame(root,text="Student's Details",font=20,bd=2,width=900,bg=framebg, fg=framefg, height=250,relief=GROOVE)
obj.place(x=30,y=200)

Label(obj,text="Full Name:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=50)
Label(obj,text="Date of Birth:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=100)
Label(obj,text="Gender:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=150)
# Label(obj,text="Gender:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=150)

Label(obj,text="Class:",font="arial 13",bg=framebg,fg=framefg).place(x=350,y=50)
Label(obj,text="Address:",font="arial 13",bg=framebg,fg=framefg).place(x=350,y=100)
Label(obj,text="Phone No:",font="arial 13",bg=framebg,fg=framefg).place(x=350,y=150)
# Label(obj,text="Phone No:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=150)

Name = StringVar()
name_entry = Entry(obj, textvariable=Name,width=20,font="arial 10")
name_entry.place(x=160,y=50)

DOB = StringVar()
dob_entry = Entry(obj, textvariable=DOB,width=20,font="arial 10")
dob_entry.place(x=160,y=100)

radio = IntVar()
R1 = Radiobutton(obj,text="Male",variable=radio,value=1,bg=framebg,fg=framefg,command=selection)
R1.place(x=150,y=150)

R2 = Radiobutton(obj,text="Female",variable=radio,value=2,bg=framebg,fg=framefg,command=selection)
R2.place(x=200,y=150)

# Class = StringVar()
# class_entry = Entry(obj, textvariable=Class,width=20,font="arial 10")
# class_entry.place(x=630,y=50)

Class = Combobox(obj,values=["NUR","L.K.G","U.K.G","1","2","3","4","5","6","7","8","9","10","11","12","Other"],font="Roboto 10",width=17,state="r")
Class.place(x=480,y=50)
Class.set("Select Class")

Address = StringVar()
address_entry = Entry(obj, textvariable=Address,width=20,font="arial 10")
address_entry.place(x=480,y=100)

Phone = StringVar()
phone_entry = Entry(obj, textvariable=Phone,width=20,font="arial 10")
phone_entry.place(x=480,y=150)
Phone.set("91+")

# Other Details
obj2= LabelFrame(root,text="Other Details",font=20,bd=2,width=900,bg=framebg, fg=framefg, height=220,relief=GROOVE)
obj2.place(x=30,y=470)

Label(obj2,text="Father's Name:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=50)
Label(obj2,text="Mother's Name:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=100)
Label(obj2,text="Total Amount:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=150)

F_Name = StringVar()
f_entry = Entry(obj2,textvariable=F_Name,width=20,font="arial 10")
f_entry.place(x=160,y=50)

M_Name = StringVar()
m_entry = Entry(obj2,textvariable=M_Name,width=20,font="arial 10")
m_entry.place(x=160,y=100)

T_Fee_Amount = StringVar()
t_f_amount_entry = Entry(obj2,textvariable=T_Fee_Amount,width=20,font="arial 10")
t_f_amount_entry.place(x=160,y=150)
T_Fee_Amount.set("₹ ")


# Label(obj2,text="Aamount Paid:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=50)
# Label(obj2,text="Amount Due:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=100)
# Label(obj2,text="Date of Payment:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=150)

# Amount_Paid = StringVar()
# a_paid_entry = Entry(obj2,textvariable=Amount_Paid,width=20,font="arial 10")
# a_paid_entry.place(x=630,y=50)

# Amount_Due = StringVar()
# amount_due_entry = Entry(obj2,textvariable=Amount_Due,width=20,font="arial 10")
# amount_due_entry.place(x=630,y=100)

# today = date.today()
# n1 = today.strftime("%d/%m/%Y")
# # print(n1)
# Date_of_Payment = StringVar()
# d_of_payment_entry = Entry(obj2,textvariable=Date_of_Payment,width=20,font="arial 10")
# d_of_payment_entry.place(x=630,y=150)
# Date_of_Payment.set(n1)

# Image
try:
    f = Frame(obj,bd=3,bg="black",width=200,height=200,relief=GROOVE)
    f.place(x=690,y=0)

    img = PhotoImage(file="Images/student.png")
    lbl = Label(f,bg="black",image=img)
    lbl.place(x=0,y=0)
    f.bind("<Button-1>", showimage)
    lbl.bind("<Button-1>", showimage)
except:
    pass

# Months & Date Of Payment
obj3 = LabelFrame(root,text="Months & Date Of Payment",font=20,bd=2,width=350,bg=framebg, fg=framefg, height=490,relief=GROOVE)
obj3.place(x=950,y=200)

Label(obj3,text="January:",font="arial 13",bg=framebg,fg=framefg).place(x=5,y=15)

# jan = Combobox(obj3,values=[f"Paid, Date: {paydate}","Not Paid"],font="Roboto 10",width=17,state="r")
# jan.place(x=100,y=50)
# jan.set("Payment")

january_month = Combobox(obj3,values=[f"Paid","Not Paid"],font="Roboto 10",width=8,state="r")
january_month.place(x=100,y=15)
january_month.set("Payment")

Janrs = StringVar()
jan_rs= Entry(obj3,textvariable=Janrs,width=7,font="arial 10")
jan_rs.place(x=188,y=15)
Janrs.set("₹ ")

JanDate = StringVar()
jan_date = Entry(obj3,textvariable=JanDate,width=12,font="arial 10")
jan_date.place(x=250,y=15)
JanDate.set(paydate)

Label(obj3,text="February:",font="arial 13",bg=framebg,fg=framefg).place(x=5,y=45)

february_month = Combobox(obj3,values=[f"Paid","Not Paid"],font="Roboto 10",width=8,state="r")
february_month.place(x=100,y=45)
february_month.set("Payment")

FebDate = StringVar()
feb_date = Entry(obj3,textvariable=FebDate,width=12,font="arial 10")
feb_date.place(x=250,y=45)
FebDate.set(paydate)

Febrs = StringVar()
feb_rs = Entry(obj3,textvariable=Febrs,width=7,font="arial 10")
feb_rs.place(x=188,y=45)
Febrs.set("₹ ")

# feb_date = DateEntry(obj3, width=5, background='darkblue',
#                 foreground='white', borderwidth=1)
# feb_date.pack(padx=0, pady=200)

Label(obj3,text="March:",font="arial 13",bg=framebg,fg=framefg).place(x=5,y=75)

march_month = Combobox(obj3,values=[f"Paid","Not Paid"],font="Roboto 10",width=8,state="r")
march_month.place(x=100,y=75)
march_month.set("Payment")


Marrs = StringVar()
mar_rs = Entry(obj3,textvariable=Marrs,width=7,font="arial 10")
mar_rs.place(x=188,y=75)
Marrs.set("₹ ")

MarDate = StringVar()
mar_date = Entry(obj3,textvariable=MarDate,width=12,font="arial 10")
mar_date.place(x=250,y=75)
MarDate.set(paydate)

Label(obj3,text="April:",font="arial 13",bg=framebg,fg=framefg).place(x=5,y=105)

april_month = Combobox(obj3,values=[f"Paid","Not Paid"],font="Roboto 10",width=8,state="r")
april_month.place(x=100,y=105)
april_month.set("Payment")

aprrs = StringVar()
apr_rs = Entry(obj3,textvariable=aprrs,width=7,font="arial 10")
apr_rs.place(x=188,y=105)
aprrs.set("₹ ")

ApDate = StringVar()
apr_date = Entry(obj3,textvariable=ApDate,width=12,font="arial 10")
apr_date.place(x=250,y=105)
ApDate.set(paydate)

Label(obj3,text="May:",font="arial 13",bg=framebg,fg=framefg).place(x=5,y=135)

may_month = Combobox(obj3,values=[f"Paid","Not Paid"],font="Roboto 10",width=8,state="r")
may_month.place(x=100,y=135)
may_month.set("Payment")

mayrs = StringVar()
may_rs = Entry(obj3,textvariable=mayrs,width=7,font="arial 10")
may_rs.place(x=188,y=135)
mayrs.set("₹ ")

MayDate = StringVar()
may_date = Entry(obj3,textvariable=MayDate,width=12,font="arial 10")
may_date.place(x=250,y=135)
MayDate.set(paydate)

Label(obj3,text="June:",font="arial 13",bg=framebg,fg=framefg).place(x=5,y=165)

June_month = Combobox(obj3,values=[f"Paid","Not Paid"],font="Roboto 10",width=8,state="r")
June_month.place(x=100,y=165)
June_month.set("Payment")

juners = StringVar()
june_rs = Entry(obj3,textvariable=juners,width=7,font="arial 10")
june_rs.place(x=188,y=165)
juners.set("₹ ")

JunDate = StringVar()
jun_date = Entry(obj3,textvariable=JunDate,width=12,font="arial 10")
jun_date.place(x=250,y=165)
JunDate.set(paydate)

Label(obj3,text="July:",font="arial 13",bg=framebg,fg=framefg).place(x=5,y=195)

July_month = Combobox(obj3,values=[f"Paid","Not Paid"],font="Roboto 10",width=8,state="r")
July_month.place(x=100,y=195)
July_month.set("Payment")

julrs = StringVar()
jul_rs = Entry(obj3,textvariable=julrs,width=7,font="arial 10")
jul_rs.place(x=188,y=195)
julrs.set("₹ ")

JulyDate = StringVar()
july_date = Entry(obj3,textvariable=JulyDate,width=12,font="arial 10")
july_date.place(x=250,y=195)
JulyDate.set(paydate)

Label(obj3,text="August:",font="arial 13",bg=framebg,fg=framefg).place(x=5,y=225)

August_month = Combobox(obj3,values=[f"Paid","Not Paid"],font="Roboto 10",width=8,state="r")
August_month.place(x=100,y=225)
August_month.set("Payment")

augrs = StringVar()
aug_rs = Entry(obj3,textvariable=augrs,width=7,font="arial 10")
aug_rs.place(x=188,y=225)
augrs.set("₹ ")

AugustDate = StringVar()
august_date = Entry(obj3,textvariable=AugustDate,width=12,font="arial 10")
august_date.place(x=250,y=225)
AugustDate.set(paydate)

Label(obj3,text="September:",font="arial 13",bg=framebg,fg=framefg).place(x=5,y=255)

September_month = Combobox(obj3,values=[f"Paid","Not Paid"],font="Roboto 10",width=8,state="r")
September_month.place(x=100,y=255)
September_month.set("Payment")

seprs = StringVar()
sep_rs = Entry(obj3,textvariable=seprs,width=7,font="arial 10")
sep_rs.place(x=188,y=255)
seprs.set("₹ ")

SeptemberDate = StringVar()
september_date = Entry(obj3,textvariable=SeptemberDate,width=12,font="arial 10")
september_date.place(x=250,y=255)
SeptemberDate.set(paydate)

Label(obj3,text="October:",font="arial 13",bg=framebg,fg=framefg).place(x=5,y=285)

October_month = Combobox(obj3,values=[f"Paid","Not Paid"],font="Roboto 10",width=8,state="r")
October_month.place(x=100,y=285)
October_month.set("Payment")

octrs = StringVar()
oct_rs = Entry(obj3,textvariable=octrs,width=7,font="arial 10")
oct_rs.place(x=188,y=285)
octrs.set("₹ ")

OctoberDate = StringVar()
october_date = Entry(obj3,textvariable=OctoberDate,width=12,font="arial 10")
october_date.place(x=250,y=285)
OctoberDate.set(paydate)

Label(obj3,text="November:",font="arial 13",bg=framebg,fg=framefg).place(x=5,y=315)

November_month = Combobox(obj3,values=[f"Paid","Not Paid"],font="Roboto 10",width=8,state="r")
November_month.place(x=100,y=315)
November_month.set("Payment")

novrs = StringVar()
nov_rs = Entry(obj3,textvariable=novrs,width=7,font="arial 10")
nov_rs.place(x=188,y=315)
novrs.set("₹ ")

NovemberDate = StringVar()
november_date = Entry(obj3,textvariable=NovemberDate,width=12,font="arial 10")
november_date.place(x=250,y=315)
NovemberDate.set(paydate)

Label(obj3,text="December:",font="arial 13",bg=framebg,fg=framefg).place(x=5,y=345)

December_month = Combobox(obj3,values=[f"Paid","Not Paid"],font="Roboto 10",width=8,state="r")
December_month.place(x=100,y=345)
December_month.set("Payment")

decrs = StringVar()
dec_rs = Entry(obj3,textvariable=decrs,width=7,font="arial 10")
dec_rs.place(x=188,y=345)
decrs.set("₹ ")

DecemberDate = StringVar()
december_date = Entry(obj3,textvariable=DecemberDate,width=12,font="arial 10")
december_date.place(x=250,y=345)
DecemberDate.set(paydate)

# Label(obj3,text="New Student Add:",font="arial 13",bg=framebg,fg=framefg).place(x=5,y=400)
# # All_Data_Search
# Search_Result_List = Combobox(obj3,values=all_students_data_list.all_student_data,font="arial 13 bold",width=15,height=2,state="r")
# Search_Result_List.place(x=150,y=400,height=38)
# Search_Result_List.set("All Student List")

# Buttons

# Button(root,text="Select Image",width=19,height=2,font="arial 12 bold",bg="lightblue",command=showimage).place(x=1000,y=370)

# saveButton=Button(root,text="Save",width=19,height=2,font="arial 12 bold",bg="lightgreen",command=Save)
# saveButton.place(x=1000,y=450)

# Button(root,text="Reset",width=19,height=2,font="arial 12 bold",bg="lightpink",command=Clear).place(x=1000,y=530)

# Button(root,text="Update",width=19,height=2,font="arial 12 bold",bg="red",command=Update).place(x=1000,y=610)

saveButton=Button(obj2,text="Save",width=19,height=2,font="arial 12 bold",bg="green",command=Save)
saveButton.place(x=450,y=50)

Button(obj2,text="Update",width=19,height=2,font="arial 12 bold",bg="#1EFF00",command=Update).place(x=670,y=50)

Button(obj2,text="Reset",width=19,height=2,font="arial 12 bold",bg="#ff0000",command=Clear).place(x=450,y=120)

Button(obj2,text="Exit",width=19,height=2,font="arial 12 bold",bg="#ad0000",command=Exit).place(x=670,y=120)

# data_copy(dst_path="D:\\")
data_copy(dst_path="C:\\")

if __name__=="__main__":
    root.mainloop()