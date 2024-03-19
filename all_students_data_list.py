# Registration No. | Student Name | Class | Gender | DOB | Date Of Registration	| Address | Phone Number | Father Name | Mother Name | Total Fee Amount	| Amount Paid | Amount Due | Date of Payment

import openpyxl, xlrd
from openpyxl import Workbook
import pathlib


all_student_data = []
data_list_formate_q = []
data_list_formate_r = []

def data_list_save_r(student_data_q):
    # try:
    file = openpyxl.load_workbook("Student Data/Student_Data.xlsx")
    sheet = file.active

    for row in sheet.rows:

        name=row[0]
        # print(str(name))
        reg_no_position = str(name)[14:-1]
        reg_number = str(name)[15:-1]
        if int(student_data_q)+1==int(reg_number):
            x1 = sheet.cell(row=int(reg_number),column=1).value
            x2 = sheet.cell(row=int(reg_number),column=2).value
            x3 = sheet.cell(row=int(reg_number),column=3).value
            x4 = sheet.cell(row=int(reg_number),column=4).value
            x5 = sheet.cell(row=int(reg_number),column=5).value
            x6 = sheet.cell(row=int(reg_number),column=6).value
            x7 = sheet.cell(row=int(reg_number),column=7).value
            x8 = sheet.cell(row=int(reg_number),column=8).value
            x9 = sheet.cell(row=int(reg_number),column=9).value
            x10 = sheet.cell(row=int(reg_number),column=10).value
            x11 = sheet.cell(row=int(reg_number),column=11).value
            x12 = sheet.cell(row=int(reg_number),column=12).value

            data_list_formate_r.append(str(f"{x1} | {x2} | {x3} | {x4} | {x5} | {x6} | {x7} | {x8} | {x9} | {x10} | {x11} | {x12}"))

def data_list_save_q():
    # try:
    file = openpyxl.load_workbook("Student Data/Student_Data.xlsx")
    sheet = file.active

    for row in sheet.rows:

        name=row[0]
        # print(str(name))
        reg_no_position = str(name)[14:-1]
        reg_number = str(name)[15:-1]
        if 1==int(reg_number):
            x1 = sheet.cell(row=int("1"),column=1).value
            x2 = sheet.cell(row=int("1"),column=2).value
            x3 = sheet.cell(row=int("1"),column=3).value
            x4 = sheet.cell(row=int("1"),column=4).value
            x5 = sheet.cell(row=int("1"),column=5).value
            x6 = sheet.cell(row=int("1"),column=6).value
            x7 = sheet.cell(row=int("1"),column=7).value
            x8 = sheet.cell(row=int("1"),column=8).value
            x9 = sheet.cell(row=int("1"),column=9).value
            x10 = sheet.cell(row=int("1"),column=10).value
            x11 = sheet.cell(row=int("1"),column=11).value
            x12 = sheet.cell(row=int("1"),column=12).value
            data_list_formate_q.append(str(f"{x1} | {x2} | {x3} | {x4} | {x5} | {x6} | {x7} | {x8} | {x9} | {x10} | {x11} | {x12}"))
       

def list_data():
    try:
        file = openpyxl.load_workbook("Student Data/Student_Data.xlsx")
        sheet = file.active

        for row in sheet.rows:

            name=row[0]
            # print(str(name))
            reg_no_position = str(name)[14:-1]
            reg_number = str(name)[15:-1]
            if reg_number=="1":
                pass
            else:
                x1 = sheet.cell(row=int(reg_number),column=1).value
                x2 = sheet.cell(row=int(reg_number),column=2).value
                x3 = sheet.cell(row=int(reg_number),column=3).value
                all_student_data.append(str(f"{x1}.> {x2} Class: {x3}"))
                # print(reg_no_position)
    except:
        pass    

list_data()

def ref_data():
    all_student_data.clear()
    try:
        file = openpyxl.load_workbook("Student Data/Student_Data.xlsx")
        sheet = file.active

        for row in sheet.rows:

            name=row[0]
            # print(str(name))
            reg_no_position = str(name)[14:-1]
            reg_number = str(name)[15:-1]
            if reg_number=="1":
                pass
            else:
                x1 = sheet.cell(row=int(reg_number),column=1).value
                x2 = sheet.cell(row=int(reg_number),column=2).value
                x3 = sheet.cell(row=int(reg_number),column=3).value
                # all_student_data.append("Save")
                all_student_data.append(str(f"{x1}.> {x2} Class: {x3}"))
                # print(reg_no_position)
    except:
        pass    

def search_data_in_option(query):
    query = query.lower()
    all_student_data.clear()
    try:
        file = openpyxl.load_workbook("Student Data/Student_Data.xlsx")
        sheet = file.active

        for row in sheet.rows:

            name=row[0]
            # print(str(name))
            reg_no_position = str(name)[14:-1]
            reg_number = str(name)[15:-1]
            if reg_number=="1":
                pass
            else:
                if str(query).lower()==str(row[0].value).lower():
                    name=row[1]
                    # print(str(name))
                    reg_no_position = str(name)[14:-1]
                    reg_number = str(name)[15:-1]
                    x1 = sheet.cell(row=int(reg_number),column=1).value
                    x2 = sheet.cell(row=int(reg_number),column=2).value
                    x3 = sheet.cell(row=int(reg_number),column=3).value
                    # print(f"{x1}.> {x2} Class: {x3}")
                    all_student_data.append(str(f"{x1}.> {x2} Class: {x3}"))

                elif str(query).lower() in str(row[1].value).lower():
                    name=row[1]
                    # print(str(name))
                    reg_no_position = str(name)[14:-1]
                    reg_number = str(name)[15:-1]
                    x1 = sheet.cell(row=int(reg_number),column=1).value
                    x2 = sheet.cell(row=int(reg_number),column=2).value
                    x3 = sheet.cell(row=int(reg_number),column=3).value
                    # print(f"{x1}.> {x2} Class: {x3}")
                    all_student_data.append(str(f"{x1}.> {x2} Class: {x3}"))
                else:
                    x1 = sheet.cell(row=int(reg_number),column=1).value
                    x2 = sheet.cell(row=int(reg_number),column=2).value
                    x3 = sheet.cell(row=int(reg_number),column=3).value
                    # all_student_data.append("Save")
                    all_student_data.append(str(f"{x1}.> {x2} Class: {x3}"))
                # print(reg_no_position)
                # elif str(query).lower() in "class":
                #     query = query.replace("class","").replace("-","").replace(":","")
                #     str(query).lower() in str(row[2].value).lower()
                #     name=row[1]
                #     # print(str(name))
                #     reg_no_position = str(name)[14:-1]
                #     reg_number = str(name)[15:-1]
                #     x1 = sheet.cell(row=int(reg_number),column=1).value
                #     x2 = sheet.cell(row=int(reg_number),column=2).value
                #     x3 = sheet.cell(row=int(reg_number),column=3).value
                #     # print(f"{x1}.> {x2} Class: {x3}")
                #     all_student_data.append(str(f"{x1}.> {x2} Class: {x3}"))


                # print(f"{x1}.> {x2} Class: {x3}")
                # print(reg_no_position)
                # print(reg_number)
                # x1 = sheet.cell(row=int(reg_number),column=1).value
                # x2 = sheet.cell(row=int(reg_number),column=2).value
                # x3 = sheet.cell(row=int(reg_number),column=3).value
                # # all_student_data.append("Save")
                # all_student_data.append(str(f"{x1}.> {x2} Class: {x3}"))
                # print(reg_no_position)
    except:
        pass    

# while True:
#     aa = input(">> ")
#     search_data_in_option(aa)

# print(all_student_data)
# all_student_data.append("Save")
# print(all_student_data)


        # all_student_data.append(str(f"{x1}.> {x2} Class: {x3}"))